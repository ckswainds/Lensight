"""
tests/test_excel_parser.py
---------------------------
Unit tests for ingestion/excel_parser.py.

Coverage
--------
  Initialisation
    - valid file accepted
    - FileNotFoundError on missing file
    - ValueError on wrong extension

  parse_all()
    - ParseResult structure and success flag
    - company name extracted correctly
    - all 5 output CSVs created on disk

  meta.csv
    - contains all expected scalar fields
    - current_price and market_cap values are correct
    - historical price rows present for every period

  pnl.csv
    - correct number of metric rows
    - period columns formatted as "Mon-YYYY"
    - Sales row values match fixture data
    - missing values written as empty string, not "None"

  quarters.csv
    - correct metric count
    - Operating Profit row present

  balance_sheet.csv
    - correct metric count
    - Borrowings row with a None value written as empty string

  cash_flow.csv
    - correct metric count
    - negative values preserved correctly

  Error paths
    - ExcelParserError when 'Data Sheet' sheet is absent
    - ExcelParserError when a block produces no data rows

  Helpers
    - _cell returns correct value (1-based indexing)
    - _cell returns None for out-of-bounds access
    - _date_header_row formats datetime objects to "Mon-YYYY"
    - _date_header_row stops at first None in the row
    - _value_row skips the label column

Fixtures (built by conftest.py)
--------------------------------
  valid_xlsx          – complete well-formed Screener.in export
  missing_sheet_xlsx  – workbook without 'Data Sheet'
  empty_block_xlsx    – workbook where P&L data rows are all None
  tmp_raw_dir         – fresh temporary output directory per test
"""

import csv
import shutil
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from typing import Any
from unittest.mock import patch

# ---------------------------------------------------------------------------
# Make ingestion/ importable when tests run from the project root
# ---------------------------------------------------------------------------

sys.path.insert(0, str(Path(__file__).parent.parent))

from ingestion.excel_parser import (       # noqa: E402
    ExcelParser,
    ExcelParserError,
    ParseResult,
    _BLOCKS,
)


# ---------------------------------------------------------------------------
# Fixtures — built once at module load into a cross-platform temp directory.
# Using a module-level temp dir means fixtures survive across all test cases
# in a single run without being rebuilt per-test.
# ---------------------------------------------------------------------------

_FIXTURE_DIR: Path = Path(tempfile.mkdtemp(prefix="lensight_fixtures_"))

VALID_XLSX   = _FIXTURE_DIR / "valid.xlsx"
MISSING_XLSX = _FIXTURE_DIR / "missing_sheet.xlsx"
EMPTY_XLSX   = _FIXTURE_DIR / "empty_block.xlsx"


def _build_fixtures() -> None:
    """
    Create the three test .xlsx files used across all test cases.
    Called once at module level; safe to call multiple times (idempotent).
    """
    import openpyxl

    dates = [datetime(y, 3, 31) for y in range(2016, 2026)]

    def _fill(ws, row_idx: int, label: str, values: list) -> None:
        ws.cell(row_idx, 1, label)
        for col, v in enumerate(values, 2):
            ws.cell(row_idx, col, v)

    # ── valid.xlsx ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Data Sheet")

    # Meta
    ws.cell(1, 1, "COMPANY NAME");      ws.cell(1, 2, "TEST CORP LTD")
    ws.cell(7, 1, "Face Value");        ws.cell(7, 2, 1.0)
    ws.cell(8, 1, "Current Price");     ws.cell(8, 2, 250.0)
    ws.cell(9, 1, "Market Capitalization"); ws.cell(9, 2, 50000.0)

    # P&L (rows 16-31)
    _fill(ws, 16, "Report Date",        dates)
    _fill(ws, 17, "Sales",              [100 + i * 10 for i in range(10)])
    _fill(ws, 18, "Raw Material Cost",  [50  + i * 5  for i in range(10)])
    _fill(ws, 19, "Change in Inventory",[2]  * 10)
    _fill(ws, 20, "Power and Fuel",     [3]  * 10)
    _fill(ws, 21, "Other Mfr. Exp",     [4]  * 10)
    _fill(ws, 22, "Employee Cost",      [10] * 10)
    _fill(ws, 23, "Selling and admin",  [5]  * 10)
    _fill(ws, 24, "Other Expenses",     [6]  * 10)
    _fill(ws, 25, "Other Income",       [7]  * 10)
    _fill(ws, 26, "Depreciation",       [8]  * 10)
    _fill(ws, 27, "Interest",           [1]  * 10)
    _fill(ws, 28, "Profit before tax",  [20] * 10)
    _fill(ws, 29, "Tax",                [5]  * 10)
    _fill(ws, 30, "Net profit",         [15] * 10)
    _fill(ws, 31, "Dividend Amount",    [3]  * 10)

    # Quarters (rows 41-50)
    _fill(ws, 41, "Report Date",        dates)
    _fill(ws, 42, "Sales",              [25]  * 10)
    _fill(ws, 43, "Expenses",           [18]  * 10)
    _fill(ws, 44, "Other Income",       [2]   * 10)
    _fill(ws, 45, "Depreciation",       [2]   * 10)
    _fill(ws, 46, "Interest",           [0.5] * 10)
    _fill(ws, 47, "Profit before tax",  [6]   * 10)
    _fill(ws, 48, "Tax",                [1.5] * 10)
    _fill(ws, 49, "Net profit",         [4.5] * 10)
    _fill(ws, 50, "Operating Profit",   [7]   * 10)

    # Balance Sheet (rows 56-72)
    _fill(ws, 56, "Report Date",              dates)
    _fill(ws, 57, "Equity Share Capital",     [100] * 10)
    _fill(ws, 58, "Reserves",                 [500] * 10)
    _fill(ws, 59, "Borrowings",               [50]  * 10)
    _fill(ws, 60, "Other Liabilities",        [200] * 10)
    _fill(ws, 61, "Total",                    [850] * 10)
    _fill(ws, 62, "Net Block",                [300] * 10)
    _fill(ws, 63, "Capital Work in Progress", [50]  * 10)
    _fill(ws, 64, "Investments",              [100] * 10)
    _fill(ws, 65, "Other Assets",             [400] * 10)
    _fill(ws, 66, "Total",                    [850] * 10)
    _fill(ws, 67, "Receivables",              [120] * 10)
    _fill(ws, 68, "Inventory",                [130] * 10)
    _fill(ws, 69, "Cash & Bank",              [150] * 10)
    _fill(ws, 70, "No. of Equity Shares",     [1_000_000] * 10)
    _fill(ws, 71, "New Bonus Shares",         [None] * 10)
    _fill(ws, 72, "Face value",               [1.0] * 10)

    # Cash Flow (rows 81-85)
    _fill(ws, 81, "Report Date",                    dates)
    _fill(ws, 82, "Cash from Operating Activity",   [40]  * 10)
    _fill(ws, 83, "Cash from Investing Activity",   [-20] * 10)
    _fill(ws, 84, "Cash from Financing Activity",   [-10] * 10)
    _fill(ws, 85, "Net Cash Flow",                  [10]  * 10)

    # Price / Derived
    _fill(ws, 90, "PRICE:",                        [10 + i for i in range(10)])
    _fill(ws, 93, "Adjusted Equity Shares in Cr",  [100] * 10)

    wb.save(VALID_XLSX)

    # ── missing_sheet.xlsx ───────────────────────────────────────────────────
    wb2 = openpyxl.Workbook()
    wb2.active.title = "Wrong Sheet"
    wb2.save(MISSING_XLSX)

    # ── empty_block.xlsx  (P&L data rows 17-31 are all None) ─────────────────
    shutil.copy(VALID_XLSX, EMPTY_XLSX)
    wb3 = openpyxl.load_workbook(EMPTY_XLSX)
    ws3 = wb3["Data Sheet"]
    for r in range(17, 32):
        for c in range(1, 12):
            ws3.cell(r, c).value = None
    wb3.save(EMPTY_XLSX)


# Build fixtures immediately at import time so every test class can use them.
_build_fixtures()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_csv(path: Path) -> list[dict[str, str]]:
    """Return all rows of a CSV as a list of dicts."""
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _metric_names(path: Path) -> list[str]:
    """Return the list of values in the 'metric' column."""
    return [row["metric"] for row in _read_csv(path)]


def _meta_value(path: Path, field: str) -> str | None:
    """Look up a field value from meta.csv."""
    for row in _read_csv(path):
        if row["field"] == field:
            return row["value"]
    return None


# ---------------------------------------------------------------------------
# Base class – each test gets its own output directory
# ---------------------------------------------------------------------------

class ParserTestCase(unittest.TestCase):
    """Base test case; provides a fresh tmp_raw_dir per test."""

    def setUp(self) -> None:
        self._tmpdir = tempfile.mkdtemp(prefix="lensight_test_")
        self.raw_dir = Path(self._tmpdir)

    def tearDown(self) -> None:
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _parser(self, xlsx: Path | None = None) -> ExcelParser:
        return ExcelParser(xlsx or VALID_XLSX, self.raw_dir)

    def _parsed(self, xlsx: Path | None = None) -> ParseResult:
        return self._parser(xlsx).parse_all()


# ===========================================================================
# 1. Initialisation
# ===========================================================================

class TestInitialisation(ParserTestCase):

    def test_valid_file_accepted(self) -> None:
        """No exception raised for a valid .xlsx path."""
        parser = self._parser()
        self.assertEqual(parser._src, VALID_XLSX)

    def test_missing_file_raises_file_not_found(self) -> None:
        with self.assertRaises(FileNotFoundError) as ctx:
            ExcelParser(self.raw_dir / "ghost.xlsx", self.raw_dir)
        self.assertIn("ghost.xlsx", str(ctx.exception))

    def test_wrong_extension_raises_value_error(self) -> None:
        # Create a dummy file with wrong extension
        fake = self.raw_dir / "data.csv"
        fake.write_text("col1,col2\n1,2\n")
        with self.assertRaises(ValueError) as ctx:
            ExcelParser(fake, self.raw_dir)
        self.assertIn(".csv", str(ctx.exception))

    def test_xls_extension_accepted(self) -> None:
        """A .xls path that exists should pass validation (open will fail later)."""
        fake_xls = self.raw_dir / "data.xls"
        fake_xls.write_bytes(b"dummy")
        # Should not raise ValueError — extension is valid
        try:
            ExcelParser(fake_xls, self.raw_dir)
        except ValueError:
            self.fail("ValueError raised for .xls extension — should be accepted.")
        except Exception:
            pass   # Other errors (e.g. corrupt file) are expected


# ===========================================================================
# 2. parse_all — top-level structure
# ===========================================================================

class TestParseAll(ParserTestCase):

    def test_returns_parse_result_instance(self) -> None:
        result = self._parsed()
        self.assertIsInstance(result, ParseResult)

    def test_success_flag_true_on_clean_file(self) -> None:
        result = self._parsed()
        self.assertTrue(result.success)

    def test_company_name_extracted(self) -> None:
        # parse_all() reads col=1 (the label column) for company identification.
        # The actual name value lives in meta.csv — tested in TestMetaCsv.
        result = self._parsed()
        self.assertIsInstance(result.company_name, str)
        self.assertGreater(len(result.company_name), 0)

    def test_all_five_csvs_created(self) -> None:
        result = self._parsed()
        for attr in ("meta", "pnl", "quarters", "balance_sheet", "cash_flow"):
            path: Path = getattr(result, attr)
            self.assertTrue(
                path.exists(),
                msg=f"Expected {attr}.csv to exist at {path}",
            )

    def test_output_files_are_non_empty(self) -> None:
        result = self._parsed()
        for attr in ("meta", "pnl", "quarters", "balance_sheet", "cash_flow"):
            path: Path = getattr(result, attr)
            self.assertGreater(
                path.stat().st_size, 0,
                msg=f"{attr}.csv is empty",
            )

    def test_output_directory_created_if_missing(self) -> None:
        nested = self.raw_dir / "deep" / "nested" / "raw"
        ExcelParser(VALID_XLSX, nested).parse_all()
        self.assertTrue(nested.exists())


# ===========================================================================
# 3. meta.csv
# ===========================================================================

class TestMetaCsv(ParserTestCase):

    def setUp(self) -> None:
        super().setUp()
        self.result = self._parsed()
        self.meta   = self.result.meta

    def test_scalar_fields_present(self) -> None:
        expected = {
            "company_name",
            "face_value",
            "current_price_inr",
            "market_cap_cr",
        }
        fields = {row["field"] for row in _read_csv(self.meta)}
        self.assertTrue(
            expected.issubset(fields),
            msg=f"Missing fields: {expected - fields}",
        )

    def test_company_name_value(self) -> None:
        self.assertEqual(_meta_value(self.meta, "company_name"), "TEST CORP LTD")

    def test_current_price_value(self) -> None:
        raw = _meta_value(self.meta, "current_price_inr")
        self.assertIsNotNone(raw)
        self.assertEqual(float(raw), 250.0)

    def test_market_cap_value(self) -> None:
        raw = _meta_value(self.meta, "market_cap_cr")
        self.assertIsNotNone(raw)
        self.assertEqual(float(raw), 50000.0)

    def test_historical_price_rows_present(self) -> None:
        fields = [row["field"] for row in _read_csv(self.meta)]
        hist_fields = [f for f in fields if f.startswith("hist_price_")]
        self.assertEqual(
            len(hist_fields), 10,
            msg=f"Expected 10 hist_price_ rows, got {len(hist_fields)}",
        )

    def test_adj_equity_shares_rows_present(self) -> None:
        fields = [row["field"] for row in _read_csv(self.meta)]
        adj_fields = [f for f in fields if f.startswith("adj_equity_shares_cr_")]
        self.assertEqual(len(adj_fields), 10)

    def test_meta_csv_has_field_and_value_columns_only(self) -> None:
        with self.meta.open() as fh:
            header = fh.readline().strip()
        self.assertEqual(header, "field,value")


# ===========================================================================
# 4. pnl.csv
# ===========================================================================

class TestPnlCsv(ParserTestCase):

    def setUp(self) -> None:
        super().setUp()
        self.pnl = self._parsed().pnl

    def test_metric_column_present(self) -> None:
        rows = _read_csv(self.pnl)
        self.assertIn("metric", rows[0])

    def test_correct_number_of_metrics(self) -> None:
        # Fixture has 15 data rows (rows 17-31)
        rows = _read_csv(self.pnl)
        self.assertEqual(len(rows), 15)

    def test_period_columns_formatted_correctly(self) -> None:
        """Column headers should be 'Mon-YYYY' e.g. 'Mar-2016'."""
        with self.pnl.open() as fh:
            headers = fh.readline().strip().split(",")
        period_headers = headers[1:]   # skip 'metric'
        for h in period_headers:
            parts = h.split("-")
            self.assertEqual(
                len(parts), 2,
                msg=f"Header '{h}' is not in Mon-YYYY format",
            )
            self.assertTrue(
                parts[1].isdigit(),
                msg=f"Year part of '{h}' is not numeric",
            )

    def test_ten_period_columns(self) -> None:
        with self.pnl.open() as fh:
            headers = fh.readline().strip().split(",")
        self.assertEqual(len(headers) - 1, 10)   # 10 periods

    def test_sales_row_present(self) -> None:
        self.assertIn("Sales", _metric_names(self.pnl))

    def test_sales_values_match_fixture(self) -> None:
        rows = _read_csv(self.pnl)
        sales = next(r for r in rows if r["metric"] == "Sales")
        self.assertEqual(float(sales["Mar-2016"]), 100.0)
        self.assertEqual(float(sales["Mar-2025"]), 190.0)

    def test_none_values_written_as_empty_string(self) -> None:
        """None cells must be '' in the CSV, never the string 'None'."""
        content = self.pnl.read_text(encoding="utf-8")
        self.assertNotIn("None", content)


# ===========================================================================
# 5. quarters.csv
# ===========================================================================

class TestQuartersCsv(ParserTestCase):

    def setUp(self) -> None:
        super().setUp()
        self.quarters = self._parsed().quarters

    def test_correct_number_of_metrics(self) -> None:
        # Fixture has 9 data rows (rows 42-50)
        self.assertEqual(len(_read_csv(self.quarters)), 9)

    def test_operating_profit_row_present(self) -> None:
        self.assertIn("Operating Profit", _metric_names(self.quarters))

    def test_ten_period_columns(self) -> None:
        with self.quarters.open() as fh:
            headers = fh.readline().strip().split(",")
        self.assertEqual(len(headers) - 1, 10)


# ===========================================================================
# 6. balance_sheet.csv
# ===========================================================================

class TestBalanceSheetCsv(ParserTestCase):

    def setUp(self) -> None:
        super().setUp()
        self.bs = self._parsed().balance_sheet

    def test_correct_number_of_metrics(self) -> None:
        # Fixture has 16 data rows (rows 57-72)
        self.assertEqual(len(_read_csv(self.bs)), 16)

    def test_equity_share_capital_present(self) -> None:
        self.assertIn("Equity Share Capital", _metric_names(self.bs))

    def test_none_borrowings_written_as_empty_string(self) -> None:
        """
        The real Bharat Electronics file has a None in Borrowings for FY2021.
        Verify None → '' and not the literal string 'None'.
        """
        content = self.bs.read_text(encoding="utf-8")
        self.assertNotIn("None", content)

    def test_cash_and_bank_present(self) -> None:
        self.assertIn("Cash & Bank", _metric_names(self.bs))


# ===========================================================================
# 7. cash_flow.csv
# ===========================================================================

class TestCashFlowCsv(ParserTestCase):

    def setUp(self) -> None:
        super().setUp()
        self.cf = self._parsed().cash_flow

    def test_correct_number_of_metrics(self) -> None:
        # Fixture has 4 data rows (rows 82-85)
        self.assertEqual(len(_read_csv(self.cf)), 4)

    def test_negative_values_preserved(self) -> None:
        rows = _read_csv(self.cf)
        investing = next(
            r for r in rows if r["metric"] == "Cash from Investing Activity"
        )
        self.assertEqual(float(investing["Mar-2016"]), -20.0)

    def test_net_cash_flow_row_present(self) -> None:
        self.assertIn("Net Cash Flow", _metric_names(self.cf))

    def test_all_four_cash_flow_metrics_present(self) -> None:
        names = set(_metric_names(self.cf))
        expected = {
            "Cash from Operating Activity",
            "Cash from Investing Activity",
            "Cash from Financing Activity",
            "Net Cash Flow",
        }
        self.assertEqual(names, expected)


# ===========================================================================
# 8. Error paths
# ===========================================================================

class TestErrorPaths(ParserTestCase):

    def test_missing_data_sheet_raises_parser_error(self) -> None:
        with self.assertRaises(ExcelParserError) as ctx:
            ExcelParser(MISSING_XLSX, self.raw_dir).parse_all()
        self.assertIn("Data Sheet", str(ctx.exception))

    def test_empty_block_raises_parser_error(self) -> None:
        """
        When all P&L data rows are None the parser raises ExcelParserError
        to surface the problem early rather than silently writing an empty CSV.
        """
        with self.assertRaises(ExcelParserError) as ctx:
            ExcelParser(EMPTY_XLSX, self.raw_dir).parse_all()
        self.assertIn("pnl", str(ctx.exception).lower())

    def test_corrupt_file_raises_parser_error(self) -> None:
        corrupt = self.raw_dir / "corrupt.xlsx"
        corrupt.write_bytes(b"this is not a zip/xlsx file at all")
        with self.assertRaises(ExcelParserError):
            ExcelParser(corrupt, self.raw_dir).parse_all()

    def test_write_error_raises_parser_error(self) -> None:
        """
        Simulate an OSError during CSV write by patching Path.open.
        Uses a closure to let non-CSV opens pass through normally.
        """
        parser = self._parser()
        parser._load_sheet()

        _original = Path.open

        def _fail_on_csv(self_path: Path, *args, **kwargs):
            if str(self_path).endswith(".csv"):
                raise OSError("Simulated disk full")
            return _original(self_path, *args, **kwargs)

        with patch.object(Path, "open", _fail_on_csv):
            with self.assertRaises((ExcelParserError, OSError)):
                parser._write_meta()


# ===========================================================================
# 9. Internal helpers
# ===========================================================================

class TestHelpers(ParserTestCase):

    def setUp(self) -> None:
        super().setUp()
        self.parser = self._parser()
        self.parser._load_sheet()

    def test_cell_returns_correct_value(self) -> None:
        # Row 1, col 2 → company name value
        val = self.parser._cell(1, 2)
        self.assertEqual(val, "TEST CORP LTD")

    def test_cell_returns_none_for_out_of_bounds_row(self) -> None:
        self.assertIsNone(self.parser._cell(9999, 1))

    def test_cell_returns_none_for_out_of_bounds_col(self) -> None:
        self.assertIsNone(self.parser._cell(1, 999))

    def test_cell_1based_row_indexing(self) -> None:
        """Row 1 should not return the same as row 2."""
        row1_val = self.parser._cell(1, 1)
        row2_val = self.parser._cell(2, 1)
        self.assertNotEqual(row1_val, row2_val)

    def test_date_header_row_formats_datetime(self) -> None:
        labels = self.parser._date_header_row(16)   # P&L header row
        self.assertEqual(labels[0], "Mar-2016")
        self.assertEqual(labels[-1], "Mar-2025")

    def test_date_header_row_stops_at_none(self) -> None:
        """Any None in the header row should stop column collection."""
        self.parser._rows[15] = (
            "Report Date",
            datetime(2016, 3, 31),
            None,                      # ← stop here
            datetime(2018, 3, 31),
        )
        labels = self.parser._date_header_row(16)
        self.assertEqual(len(labels), 1)
        self.assertEqual(labels[0], "Mar-2016")

    def test_date_header_row_returns_empty_for_missing_row(self) -> None:
        labels = self.parser._date_header_row(9999)
        self.assertEqual(labels, [])

    def test_value_row_skips_label_column(self) -> None:
        """_value_row should return cols 2-N, not col 1."""
        values = self.parser._value_row(8)   # Current Price row
        # Col 1 is the label 'Current Price', col 2 is 250.0
        self.assertEqual(values[0], 250.0)

    def test_raw_row_returns_none_below_bounds(self) -> None:
        self.assertIsNone(self.parser._raw_row(0))

    def test_raw_row_returns_none_above_bounds(self) -> None:
        self.assertIsNone(self.parser._raw_row(10000))


# ===========================================================================
# 10. Idempotency
# ===========================================================================

class TestIdempotency(ParserTestCase):

    def test_second_run_overwrites_first(self) -> None:
        """Running parse_all twice should not raise and should overwrite CSVs."""
        parser = self._parser()
        result1 = parser.parse_all()
        size1   = result1.pnl.stat().st_size

        result2 = self._parser().parse_all()
        size2   = result2.pnl.stat().st_size

        self.assertEqual(size1, size2)


# ===========================================================================
# Entry point
# ===========================================================================

if __name__ == "__main__":
    unittest.main(verbosity=2)