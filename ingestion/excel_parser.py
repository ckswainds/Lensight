"""
ingestion/excel_parser.py
--------------------------
Parses a Screener.in Excel export and splits it into five
raw CSV files, one per financial statement.

The Screener.in workbook stores all numerical data in a single
sheet called 'Data Sheet'. Every other sheet (Profit & Loss,
Balance Sheet, etc.) just references cells in that master sheet
via formulas. This parser therefore reads only 'Data Sheet',
using data_only=True so openpyxl resolves the cached values
instead of returning formula strings.

Output layout
-------------
data/raw/
    meta.csv           – company name, face value, price, market cap
    pnl.csv            – annual profit & loss (10 years)
    quarters.csv       – quarterly financials (10 quarters)
    balance_sheet.csv  – annual balance sheet (10 years)
    cash_flow.csv      – annual cash flow statement (10 years)

All CSVs are written in long format:
    metric | FY2016 | FY2017 | ... | FY2025

Row map (Data Sheet, 1-based)
------------------------------
 1        company name
 7        face value
 8        current price
 9        market capitalisation
16–31     profit & loss  (row 16 = header / report dates)
41–50     quarterly data (row 41 = header)
56–72     balance sheet  (row 56 = header)
81–85     cash flow      (row 81 = header)
90        historical price per share
93        adjusted equity shares in crores
"""

import csv
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Public dataclass – returned by ExcelParser.parse_all()
# ---------------------------------------------------------------------------

@dataclass
class ParseResult:
    """Holds the output paths produced by a single parse run."""
    company_name: str
    meta:          Path
    pnl:           Path
    quarters:      Path
    balance_sheet: Path
    cash_flow:     Path
    errors:        list[str] = field(default_factory=list)

    @property
    def success(self) -> bool:
        return len(self.errors) == 0


# ---------------------------------------------------------------------------
# Custom exception
# ---------------------------------------------------------------------------

class ExcelParserError(Exception):
    """Raised when the workbook cannot be parsed."""


# ---------------------------------------------------------------------------
# Row-range configuration  (1-based, inclusive on both ends)
# ---------------------------------------------------------------------------

_SHEET          = "Data Sheet"
_ROW_COMPANY    = 1
_ROW_FACE_VALUE = 7
_ROW_PRICE      = 8
_ROW_MARKET_CAP = 9
_ROW_HIST_PRICE = 90
_ROW_ADJ_SHARES = 93

_BLOCKS: dict[str, tuple[int, int]] = {
    # statement_name: (header_row, last_data_row)
    "pnl":           (16, 31),
    "quarters":      (41, 50),
    "balance_sheet": (56, 72),
    "cash_flow":     (81, 85),
}


# ---------------------------------------------------------------------------
# Main parser class
# ---------------------------------------------------------------------------

class ExcelParser:
    """
    Splits a Screener.in .xlsx export into raw CSVs.

    Parameters
    ----------
    xlsx_path : path-like
        Path to the uploaded .xlsx file.
    raw_dir : path-like
        Destination folder for raw CSVs (data/raw/).

    Raises
    ------
    FileNotFoundError
        If ``xlsx_path`` does not exist on disk.
    ValueError
        If the file extension is not .xlsx or .xls.
    ExcelParserError
        For any structural problem found inside the workbook.

    Example
    -------
    >>> parser = ExcelParser("data/uploads/Bharat_Electronics.xlsx", "data/raw")
    >>> result = parser.parse_all()
    >>> print(result.pnl)
    data/raw/pnl.csv
    """

    def __init__(self, xlsx_path: str | Path, raw_dir: str | Path) -> None:
        self._src  = Path(xlsx_path)
        self._dest = Path(raw_dir)
        self._rows: list[tuple[Any, ...]] = []

        self._validate_inputs()

    # ------------------------------------------------------------------
    # Public interface
    # ------------------------------------------------------------------

    def parse_all(self) -> ParseResult:
        """
        Run the full parse pipeline.

        Loads the workbook once, extracts every financial block,
        writes one CSV per statement, then closes the workbook.

        Returns
        -------
        ParseResult
            Dataclass with the path of each output CSV and the
            company name extracted from the file.

        Raises
        ------
        ExcelParserError
            If the workbook cannot be opened or lacks 'Data Sheet'.
        """
        logger.info("Opening workbook: %s", self._src.name)
        self._dest.mkdir(parents=True, exist_ok=True)
        self._load_sheet()

        company = self._cell(_ROW_COMPANY, col=1) or "UNKNOWN"
        logger.info("Company identified as: %s", company)

        result = ParseResult(
            company_name=str(company),
            meta=Path(),
            pnl=Path(),
            quarters=Path(),
            balance_sheet=Path(),
            cash_flow=Path(),
        )

        result.meta          = self._write_meta()
        result.pnl           = self._write_block("pnl")
        result.quarters      = self._write_block("quarters")
        result.balance_sheet = self._write_block("balance_sheet")
        result.cash_flow     = self._write_block("cash_flow")

        logger.info(
            "Parse complete — 5 CSVs written to '%s'", self._dest
        )
        return result

    # ------------------------------------------------------------------
    # Workbook loading
    # ------------------------------------------------------------------

    def _load_sheet(self) -> None:
        """
        Open the workbook in read-only / data-only mode and cache
        every row from 'Data Sheet' as a list of value tuples.

        data_only=True instructs openpyxl to return the last
        calculated value for formula cells rather than the formula
        string itself.
        """
        try:
            wb = openpyxl.load_workbook(
                self._src, data_only=True, read_only=True
            )
        except Exception as exc:
            raise ExcelParserError(
                f"Cannot open '{self._src.name}': {exc}"
            ) from exc

        if _SHEET not in wb.sheetnames:
            wb.close()
            raise ExcelParserError(
                f"Sheet '{_SHEET}' not found. "
                f"Available sheets: {wb.sheetnames}. "
                "Ensure the file is a valid Screener.in export."
            )

        ws = wb[_SHEET]
        self._rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

        logger.debug(
            "Cached %d rows from '%s'", len(self._rows), _SHEET
        )

    # ------------------------------------------------------------------
    # Meta extraction
    # ------------------------------------------------------------------

    def _write_meta(self) -> Path:
        """
        Extract scalar company metadata and write meta.csv.

        Fields captured:
            company_name, face_value, current_price_inr,
            market_cap_cr, hist_price_series, adj_equity_shares_cr
        """
        logger.debug("Extracting meta block...")

        price_dates = self._date_header_row(_BLOCKS["pnl"][0])
        hist_prices = self._value_row(_ROW_HIST_PRICE)
        adj_shares  = self._value_row(_ROW_ADJ_SHARES)

        # Col-0 (1-based: col=1) is the row label; col-1 (1-based: col=2) is the value
        scalar_rows: list[dict[str, Any]] = [
            {"field": "company_name",      "value": self._cell(_ROW_COMPANY, 2)},
            {"field": "face_value",        "value": self._cell(_ROW_FACE_VALUE, 2)},
            {"field": "current_price_inr", "value": self._cell(_ROW_PRICE, 2)},
            {"field": "market_cap_cr",     "value": self._cell(_ROW_MARKET_CAP, 2)},
        ]

        # Append one row per year for price and adjusted shares
        for date_label, price, shares in zip(price_dates, hist_prices, adj_shares):
            scalar_rows.append({
                "field": f"hist_price_{date_label}",
                "value": price,
            })
            scalar_rows.append({
                "field": f"adj_equity_shares_cr_{date_label}",
                "value": shares,
            })

        out = self._dest / "meta.csv"
        self._write_csv(out, fieldnames=["field", "value"], rows=scalar_rows)
        logger.info("meta.csv  → %d fields", len(scalar_rows))
        return out

    # ------------------------------------------------------------------
    # Generic block writer
    # ------------------------------------------------------------------

    def _write_block(self, name: str) -> Path:
        """
        Extract a financial statement block and write it as a CSV.

        The first row of each block contains the report dates (column
        headers). Subsequent rows are metric rows: col-0 is the metric
        name, cols 1-N are the period values.

        Parameters
        ----------
        name : str
            Key from ``_BLOCKS`` — one of pnl, quarters,
            balance_sheet, cash_flow.

        Returns
        -------
        Path
            Path of the written CSV file.
        """
        header_row, last_row = _BLOCKS[name]
        logger.debug(
            "Extracting block '%s' (rows %d–%d)...",
            name, header_row, last_row,
        )

        date_labels = self._date_header_row(header_row)
        if not date_labels:
            raise ExcelParserError(
                f"No period headers found for block '{name}' "
                f"at row {header_row}."
            )

        fieldnames = ["metric"] + date_labels

        data_rows: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, last_row + 1):
            raw = self._raw_row(row_idx)
            if raw is None or not any(v is not None for v in raw):
                continue   # skip fully empty rows silently

            metric = raw[0]
            if metric is None:
                continue

            values = list(raw[1 : len(date_labels) + 1])
            # Pad with None if fewer values than headers
            values += [None] * (len(date_labels) - len(values))

            row_dict: dict[str, Any] = {"metric": str(metric).strip()}
            for label, val in zip(date_labels, values):
                row_dict[label] = val if val is not None else ""
            data_rows.append(row_dict)

        if not data_rows:
            raise ExcelParserError(
                f"Block '{name}' produced no data rows "
                f"(rows {header_row}–{last_row})."
            )

        out = self._dest / f"{name}.csv"
        self._write_csv(out, fieldnames=fieldnames, rows=data_rows)
        logger.info(
            "%-14s → %d metrics × %d periods",
            f"{name}.csv", len(data_rows), len(date_labels),
        )
        return out

    # ------------------------------------------------------------------
    # Low-level helpers
    # ------------------------------------------------------------------

    def _raw_row(self, row_num: int) -> tuple[Any, ...] | None:
        """Return the raw tuple for a 1-based row index, or None."""
        idx = row_num - 1
        if idx < 0 or idx >= len(self._rows):
            return None
        return self._rows[idx]

    def _cell(self, row_num: int, col: int) -> Any:
        """Return a single cell value (1-based row and col)."""
        row = self._raw_row(row_num)
        if row is None or col - 1 >= len(row):
            return None
        return row[col - 1]

    def _date_header_row(self, row_num: int) -> list[str]:
        """
        Read the period header row and return a list of formatted
        date strings (e.g. "Mar-2016").  The first column is the
        metric-name column and is skipped.
        """
        raw = self._raw_row(row_num)
        if raw is None:
            return []

        labels: list[str] = []
        for val in raw[1:]:
            if val is None:
                break
            if isinstance(val, datetime):
                labels.append(val.strftime("%b-%Y"))
            else:
                labels.append(str(val).strip())
        return labels

    def _value_row(self, row_num: int) -> list[Any]:
        """
        Return all non-label values from a row (skips col-0).
        """
        raw = self._raw_row(row_num)
        if raw is None:
            return []
        return list(raw[1:])

    # ------------------------------------------------------------------
    # CSV writer
    # ------------------------------------------------------------------

    @staticmethod
    def _write_csv(
        path: Path,
        fieldnames: list[str],
        rows: list[dict[str, Any]],
    ) -> None:
        """
        Write ``rows`` to ``path`` as a UTF-8 CSV.

        Uses csv.DictWriter so column ordering is deterministic.
        extrasaction='ignore' silently drops any key not in
        ``fieldnames`` — prevents crashes from unexpected extra
        columns in future Screener.in format updates.
        """
        try:
            with path.open("w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(
                    fh,
                    fieldnames=fieldnames,
                    extrasaction="ignore",
                )
                writer.writeheader()
                writer.writerows(rows)
        except OSError as exc:
            raise ExcelParserError(
                f"Failed to write '{path}': {exc}"
            ) from exc

    # ------------------------------------------------------------------
    # Input validation
    # ------------------------------------------------------------------

    def _validate_inputs(self) -> None:
        if not self._src.exists():
            raise FileNotFoundError(
                f"Upload not found: '{self._src}'. "
                "Place the .xlsx file in data/uploads/ before parsing."
            )
        if self._src.suffix.lower() not in {".xlsx", ".xls"}:
            raise ValueError(
                f"Expected a .xlsx or .xls file, got '{self._src.suffix}'. "
                "Only Screener.in Excel exports are supported."
            )